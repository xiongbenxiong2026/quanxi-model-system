"""
全犀模型系统 - 数据导出到Excel（真实业务数据版）
2026年6月真实指标：
  直播间53个 | 品牌方540家 | 总销售额1.344亿 | 8个流量方
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random, os, datetime

random.seed(42)

# ====== 真实指标 ======
TOTAL_SALES = 134400000      # 1亿3440万
BRAND_SALES = 82000000       # 8200万 (62%)
OTHER_SALES = TOTAL_SALES - BRAND_SALES  # 5240万 (38%)
TOTAL_AUDIENCE = 636000      # 63.6万人
STUDIO_COUNT = 53
BRAND_COUNT = 540
TRAFFIC_COUNT = 8

CITIES = [
    '北京市', '上海市', '广州市', '深圳市', '杭州市', '成都市', '武汉市',
    '南京市', '重庆市', '天津市', '苏州市', '西安市', '长沙市', '郑州市',
    '东莞市', '青岛市', '合肥市', '佛山市', '宁波市', '昆明市', '沈阳市',
    '大连市', '厦门市', '济南市', '南宁市', '太原市', '贵阳市', '南昌市'
]
BRAND_CATEGORIES = [
    '食品零食', '美妆护肤', '家居日用', '服装鞋包', '母婴用品',
    '数码家电', '健康保健', '宠物用品', '运动户外', '珠宝饰品'
]

def randint(a, b): return random.randint(a, b)
def randfloat(a, b, d=2): return round(random.uniform(a, b), d)
def pick(arr): return random.choice(arr)
def gdate(start='2026-05-01', days=38):
    d = datetime.date(2026, 5, 1) + datetime.timedelta(days=randint(0, days))
    return d.isoformat()

# ====== 生成直播间（53个，合计636000人） ======
def generate_studios():
    studios = []
    sizes = []
    for _ in range(30): sizes.append(randint(4000, 10000))
    for _ in range(10): sizes.append(randint(10000, 20000))
    for _ in range(8): sizes.append(randint(20000, 35000))
    for _ in range(5): sizes.append(randint(35000, 80000))

    cur = sum(sizes)
    diff = TOTAL_AUDIENCE - cur
    for _ in range(abs(diff)):
        idx = randint(0, len(sizes)-1)
        sizes[idx] += 1 if diff > 0 else -1
        sizes[idx] = max(4000, min(80000, sizes[idx]))

    traffic_types = ['线上', '线下', '混合']
    for i in range(len(sizes)):
        tt = pick(traffic_types)
        offline = tt in ('线下', '混合')
        audience = sizes[i]
        ratio = audience / TOTAL_AUDIENCE
        amount = round(TOTAL_SALES * ratio, 2)
        refund = round(amount * randfloat(0.03, 0.08), 2)

        studios.append({
            '编号': f'ZB{str(i+1).zfill(3)}',
            '名称': '直播间 ***',
            '流量类型': tt,
            '地点': pick(CITIES),
            '人数': audience,
            '门店数(线下)': randint(5, 200) if offline else 0,
            '订单金额(元)': amount,
            '退货金额(元)': refund,
            '退货率(%)': round(refund/amount*100, 1),
            '日均客单价(元)': 7,
            '每天直播时长(小时)': 1,
            '状态': '活跃'
        })
    return studios

# ====== 生成品牌方（540家） ======
def generate_brands():
    brands = []
    types = ['高品质', '热销爆款', '新品', '经典款']
    forms = ['系列', '套装', '礼盒', '单品']
    settlements = ['T+1', '周结', '月结']
    statuses = ['活跃', '活跃', '活跃', '待审', '活跃']

    for i in range(BRAND_COUNT):
        if i < 10:
            ms = randint(2000000, 8000000)
        elif i < 50:
            ms = randint(300000, 2000000)
        elif i < 200:
            ms = randint(50000, 300000)
        else:
            ms = randint(5000, 50000)

        brands.append({
            '编号': f'PP{str(i+1).zfill(4)}',
            '名称': '品牌 ***',
            '品类': pick(BRAND_CATEGORIES),
            '产品信息': f'{pick(types)}{pick(forms)}',
            '价格带(元)': f'{randint(19, 99)}-{randint(100, 999)}',
            '月销售额(元)': ms,
            '结算方式': pick(settlements),
            '入驻日期': gdate(),
            '状态': pick(statuses)
        })
    brands.sort(key=lambda x: x['月销售额(元)'], reverse=True)
    return brands

# ====== 生成订单（运营期30天样本） ======
def generate_orders(studios, brands):
    orders = []
    payments = ['微信支付', '支付宝', '银联']
    statuses = ['已完成', '已完成', '已完成', '已完成', '退款中', '已退款']
    idx = 0

    for day in range(30):
        d = datetime.date(2026, 5, 1) + datetime.timedelta(days=day)
        ds = d.isoformat()

        for _ in range(200):
            studio = pick(studios)
            brand = pick(brands)
            qty = randint(1, 5)
            price = randfloat(10, 199)
            amount = round(price * qty, 2)
            refund = round(amount * randfloat(0.3, 1), 2) if random.random() < 0.06 else 0
            idx += 1

            orders.append({
                '订单号': f'QX{ds.replace("-","")}{str(idx).zfill(6)}',
                '直播间': studio['名称'],
                '品牌方': brand['名称'],
                '品牌品类': brand['品类'],
                '金额(元)': amount,
                '数量': qty,
                '退款金额(元)': refund,
                '流量来源': pick(['线上（云流量）', '线上（云流量）', '线上（云流量）', '线下（实体流量）']),
                '日期': ds,
                '状态': pick(statuses),
                '支付方式': pick(payments)
            })
    orders.sort(key=lambda x: x['日期'])
    return orders

# ====== 生成流量方（8个） ======
def generate_traffic():
    sources = [
        ('老庞服务商', '线下拉新', '郑州本地社区团长网络'),
        ('徐阳渠道', '线下拉新', '杭州区域商超合作渠道'),
        ('成都地推团队', '线下拉新', '成都及周边地推网络'),
        ('广州批发渠道', '线下拉新', '广州批发市场商户联盟'),
        ('精准粉团队', '打粉团队', '信息流投放精准获客'),
        ('短视频引流组', '打粉团队', '抖音/视频号矩阵引流'),
        ('社群裂变组', '打粉团队', '微信社群裂变运营'),
        ('直播切片分发', '打粉团队', '直播精彩片段多平台分发')
    ]
    result = []
    for i, (name, tp, desc) in enumerate(sources):
        result.append({
            '名称': name,
            '类型': tp,
            '地区': pick(CITIES),
            '覆盖人数': randint(8000, 80000),
            '月产能(人)': randint(15000, 150000),
            '佣金比例(%)': randfloat(8, 25),
            '合作日期': gdate('2026-04-01', 40),
            '状态': '合作中',
            '说明': desc
        })
    return result

# ====== Excel 导出 ======
def style_header(ws):
    hf = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='2563EB', end_color='1D4ED8', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    for cell in ws[1]:
        cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = tb

def auto_width(ws, mx=22):
    for col_cells in ws.columns:
        ml = 0
        cl = get_column_letter(col_cells[0].column)
        for c in col_cells:
            if c.value:
                l = sum(2 if ord(ch) > 127 else 1 for ch in str(c.value))
                ml = max(ml, l)
        ws.column_dimensions[cl].width = min(max(ml + 4, 10), mx)

def export_excel(data_list, filename, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if not data_list:
        wb.save(filename); return

    headers = list(data_list[0].keys())
    ws.append(headers)
    style_header(ws)

    df = Font(name='微软雅黑', size=10)
    da = Alignment(horizontal='center', vertical='center')
    tb = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    lf = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    for rd in data_list:
        ws.append([rd[h] for h in headers])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = df; cell.alignment = da; cell.border = tb

    for r in range(2, ws.max_row + 1, 2):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = lf

    auto_width(ws)
    wb.save(filename)
    print(f'[OK] {os.path.basename(filename)} ({ws.max_row - 1} 条)')

def main():
    out = os.path.expanduser('~/Downloads')
    print('生成数据...')
    studios = generate_studios()
    brands = generate_brands()
    orders = generate_orders(studios, brands)
    traffic = generate_traffic()
    print('导出 Excel...')
    export_excel(studios, os.path.join(out, '全犀模型_01_直播间数据.xlsx'), '直播间')
    export_excel(brands, os.path.join(out, '全犀模型_02_品牌方数据.xlsx'), '品牌方')
    export_excel(orders, os.path.join(out, '全犀模型_03_订单数据.xlsx'), '订单')
    export_excel(traffic, os.path.join(out, '全犀模型_04_流量方数据.xlsx'), '流量方')
    print(f'\n[OK] 全部导出完成！保存在: {out}')

if __name__ == '__main__':
    main()
